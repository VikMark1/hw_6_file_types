import codecs
import os
import shutil
from zipfile import ZipFile
import zipfile
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import csv
import pytest


def test_create_archive():
    archive = zipfile.ZipFile(os.path.abspath('archive.zip'), 'w')
    archive.write('./resources/client_data.xlsx')
    archive.write('./resources/example1.csv')
    archive.write('./resources/sample1.pdf')
    archive.close()

def test_move_archive():
    src_path = os.path.abspath('archive.zip')
    dst_path = os.path.abspath('./resources')
    shutil.move(src_path, dst_path)

def test_files_list():
    zip1 = zipfile.ZipFile('./resources/archive.zip', 'r')
    print(zip1.namelist())
    zip1.close()

def test_check_content_xlsx():
    with zipfile.ZipFile(os.path.abspath("./resources/archive.zip")) as myzip:
        with myzip.open('resources/client_data.xlsx') as xlfile:
            xlfile = load_workbook(xlfile)
            sheet = xlfile.active
            data = sheet.cell(row=3, column=2).value
            assert data == "Mara"

def test_check_content_pdf():
    with zipfile.ZipFile(os.path.abspath("./resources/archive.zip")) as myzip1:
        with myzip1.open('resources/sample1.pdf') as pdfFile:
            pdfFile = PdfReader(pdfFile)
            number_of_pages = len(pdfFile.pages)
            assert number_of_pages == 1


def test_check_content_csv():
    with zipfile.ZipFile(os.path.abspath("./resources/archive.zip")) as myzip2:
        with myzip2.open('resources/example1.csv') as csvFile:
            table = csv.reader(codecs.iterdecode(csvFile, 'utf-8'))
            for line_no, line in enumerate(table, 1):
                if line_no == 2:
                    assert line[1] == '2016-01-01'

